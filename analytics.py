import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import plotly.express as px
except ModuleNotFoundError:
    px = None

from config import (
    ERROR_LOG_FILE,
    INPUT_FILE,
    OUTPUT_DIR,
    REPORT_FILE,
    REQUIRED_COLUMNS,
    SUMMARY_FILE,
)


QUARTER_PATTERN = re.compile(r"^Q[1-4]\s20\d{2}$")
NAVY = "#244a73"
BLUE = "#2f6f9f"
TEAL = "#248f8f"
GOLD = "#c99700"
RED = "#b91c1c"
GREEN = "#147d64"


def log_error(message):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(ERROR_LOG_FILE, "a", encoding="utf-8") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")


def load_data(file_path=INPUT_FILE):
    try:
        if not Path(file_path).exists():
            log_error(f"Input file not found: {file_path}")
            return pd.DataFrame(), [f"Input file not found: {file_path}"]
        return pd.read_excel(file_path), []
    except Exception as exc:
        log_error(f"Failed to load workbook: {exc}")
        return pd.DataFrame(), [f"Failed to load workbook: {exc}"]


def load_uploaded_data(uploaded_file):
    try:
        if uploaded_file is None:
            return pd.DataFrame(), ["No uploaded workbook was provided."]
        return pd.read_excel(uploaded_file), []
    except Exception as exc:
        log_error(f"Failed to load uploaded workbook: {exc}")
        return pd.DataFrame(), [f"Failed to load uploaded workbook: {exc}"]


def validate_data(df):
    errors = []
    warnings = []
    if df.empty:
        errors.append("Input workbook is empty.")
        log_error("Input workbook is empty.")
        return df, errors, warnings

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in df.columns]
    if missing_columns:
        message = f"Missing required columns: {', '.join(missing_columns)}"
        errors.append(message)
        log_error(message)
        return df, errors, warnings

    valid_mask = pd.Series(True, index=df.index)
    numeric_rules = {
        "participants": (0, None),
        "completion_rate": (0, 100),
        "job_placement_rate": (0, 100),
        "funding_amount": (0, None),
    }

    for column, (minimum, maximum) in numeric_rules.items():
        values = pd.to_numeric(df[column], errors="coerce")
        invalid = values.isna() | (values < minimum)
        if maximum is not None:
            invalid = invalid | (values > maximum)
        if invalid.any():
            invalid_rows = df.index[invalid].tolist()
            warnings.append(f"Skipped {len(invalid_rows)} row(s) with invalid {column}.")
            log_error(f"Invalid {column} in row indexes: {invalid_rows}")
            valid_mask &= ~invalid

    quarter_invalid = ~df["quarter"].astype(str).str.strip().str.match(QUARTER_PATTERN)
    if quarter_invalid.any():
        invalid_rows = df.index[quarter_invalid].tolist()
        warnings.append(f"Skipped {len(invalid_rows)} row(s) with unsupported quarter format.")
        log_error(f"Invalid quarter format in row indexes: {invalid_rows}")
        valid_mask &= ~quarter_invalid

    return df.loc[valid_mask].copy(), errors, warnings


def clean_data(df):
    cleaned = df.copy()
    numeric_defaults = {
        "participants": 0,
        "completion_rate": 0,
        "job_placement_rate": 0,
        "funding_amount": 0,
        "target_completion_rate": 80,
        "target_job_placement_rate": 75,
        "operating_cost": 0,
        "maritime_training_hours": 0,
        "internship_pipeline_count": 0,
        "employer_partnerships": 0,
        "credential_earned_rate": 0,
        "port_related_job_placements": 0,
    }
    text_defaults = {
        "sector_focus": "Transportation, Logistics, and Distribution",
        "county": "Regional",
        "demographic_focus": "Regional Workforce",
        "logistics_specialization": "Workforce Pipeline Development",
    }

    for column, default in {**numeric_defaults, **text_defaults}.items():
        if column not in cleaned.columns:
            cleaned[column] = default

    for column, default in numeric_defaults.items():
        cleaned[column] = pd.to_numeric(cleaned[column], errors="coerce").fillna(default)

    for column, default in text_defaults.items():
        cleaned[column] = cleaned[column].fillna(default).astype(str).str.strip()
        cleaned.loc[cleaned[column] == "", column] = default

    cleaned["quarter"] = cleaned["quarter"].astype(str).str.strip()
    cleaned["quarter_sort"] = cleaned["quarter"].apply(_quarter_sort_key)
    cleaned["funding_efficiency"] = np.where(
        cleaned["participants"] > 0,
        cleaned["funding_amount"] / cleaned["participants"],
        0,
    )
    cleaned["completion_count"] = cleaned["participants"] * cleaned["completion_rate"] / 100
    cleaned["placement_count"] = cleaned["participants"] * cleaned["job_placement_rate"] / 100
    cleaned["participant_growth_score"] = _participant_growth_score(cleaned)
    cleaned["workforce_readiness_score"] = _workforce_readiness(cleaned)
    cleaned["participant_retention"] = cleaned["completion_rate"]
    cleaned["program_label"] = cleaned.apply(_program_label, axis=1)
    return cleaned.sort_values(["quarter_sort", "program_name"]).reset_index(drop=True)


def calculate_kpis(df):
    if df.empty:
        return {}
    total_participants = float(df["participants"].sum())
    maritime_df = df[
        df["logistics_specialization"].str.contains("maritime|port|terminal|cargo", case=False, na=False)
    ]
    maritime_placement = (
        maritime_df["job_placement_rate"].mean()
        if not maritime_df.empty
        else df["job_placement_rate"].mean()
    )
    return {
        "total_participants": int(total_participants),
        "avg_completion_rate": round(df["completion_rate"].mean(), 1),
        "avg_job_placement_rate": round(df["job_placement_rate"].mean(), 1),
        "total_funding": round(df["funding_amount"].sum(), 2),
        "active_programs": int(df["program_name"].nunique()),
        "maritime_placement_rate": round(maritime_placement, 1),
        "avg_readiness_score": round(df["workforce_readiness_score"].mean(), 1),
        "avg_funding_efficiency": round(df["funding_efficiency"].replace([np.inf, -np.inf], 0).mean(), 2),
    }


def prepare_display_tables(trends, top_programs, bottom_programs, funding):
    return {
        "trends": _rename_for_display(
            trends.drop(columns=["quarter_sort"], errors="ignore"),
            {
                "quarter": "Reporting Quarter",
                "participants": "Workforce Participants",
                "completion_rate": "Avg Completion Rate (%)",
                "job_placement_rate": "Avg TLD Placement Rate (%)",
                "credential_earned_rate": "Credential Attainment (%)",
                "funding_amount": "Funding Allocation ($)",
                "employer_partnerships": "Employer Partnerships",
                "workforce_readiness_score": "Readiness Score",
                "participant_growth": "Participant Growth (%)",
                "workforce_completions": "Estimated Completions",
                "placement_growth": "Placement Change (pts)",
                "funding_growth": "Funding Growth (%)",
                "trend_classification": "Trend Signal",
            },
        ),
        "top_programs": _rename_for_display(top_programs, _program_display_columns()),
        "bottom_programs": _rename_for_display(bottom_programs, _program_display_columns()),
        "funding": _rename_for_display(
            funding,
            {
                "program_name": "Workforce Initiative",
                "participants": "Participants Served",
                "funding_amount": "Funding Allocation ($)",
                "job_placement_rate": "TLD Placement Rate (%)",
                "workforce_readiness_score": "Readiness Score",
                "funding_efficiency": "Cost per Participant ($)",
                "estimated_cost_per_placement": "Estimated Cost per Placement ($)",
            },
        ),
    }


def calculate_trends(df):
    if df.empty:
        return pd.DataFrame()
    trend_df = (
        df.groupby(["quarter", "quarter_sort"], as_index=False)
        .agg(
            participants=("participants", "sum"),
            completion_rate=("completion_rate", "mean"),
            job_placement_rate=("job_placement_rate", "mean"),
            credential_earned_rate=("credential_earned_rate", "mean"),
            funding_amount=("funding_amount", "sum"),
            employer_partnerships=("employer_partnerships", "sum"),
            workforce_readiness_score=("workforce_readiness_score", "mean"),
        )
        .sort_values("quarter_sort")
    )
    trend_df["participant_growth"] = trend_df["participants"].pct_change().replace([np.inf, -np.inf], 0).fillna(0) * 100
    trend_df["workforce_completions"] = trend_df["participants"] * trend_df["completion_rate"] / 100
    trend_df["placement_growth"] = trend_df["job_placement_rate"].diff().fillna(0)
    trend_df["funding_growth"] = trend_df["funding_amount"].pct_change().replace([np.inf, -np.inf], 0).fillna(0) * 100
    trend_df["trend_classification"] = trend_df["participant_growth"].apply(_classify_trend)
    return trend_df


def identify_top_programs(df, n=5):
    if df.empty:
        return pd.DataFrame()
    ranked = _program_rankings(df)
    return ranked.head(n)


def identify_bottom_programs(df, n=5):
    if df.empty:
        return pd.DataFrame()
    ranked = _program_rankings(df)
    return ranked.tail(n).sort_values("ranking_score")


def calculate_funding_efficiency(df):
    if df.empty:
        return pd.DataFrame()
    funding = (
        df.groupby("program_name", as_index=False)
        .agg(
            participants=("participants", "sum"),
            funding_amount=("funding_amount", "sum"),
            job_placement_rate=("job_placement_rate", "mean"),
            workforce_readiness_score=("workforce_readiness_score", "mean"),
        )
    )
    funding["funding_efficiency"] = np.where(
        funding["participants"] > 0,
        funding["funding_amount"] / funding["participants"],
        0,
    )
    funding["estimated_cost_per_placement"] = np.where(
        funding["participants"] * funding["job_placement_rate"] / 100 > 0,
        funding["funding_amount"] / (funding["participants"] * funding["job_placement_rate"] / 100),
        0,
    )
    return funding.sort_values("funding_efficiency")


def generate_visualizations(df):
    figures = {}
    if df.empty:
        return figures
    if px is None:
        log_error("Plotly is not installed. Install requirements.txt before rendering dashboard visualizations.")
        return figures

    trends = calculate_trends(df)
    top = identify_top_programs(df)
    bottom = identify_bottom_programs(df)
    county = df.groupby("county", as_index=False)["participants"].sum().sort_values("participants", ascending=False)
    specialization = df.groupby("logistics_specialization", as_index=False)["participants"].sum()
    heatmap_data = df.pivot_table(
        index="program_name",
        columns="quarter",
        values="workforce_readiness_score",
        aggfunc="mean",
    )
    heatmap_data = heatmap_data.reindex(sorted(heatmap_data.columns, key=_quarter_sort_key), axis=1)
    program_rates = (
        df.groupby("program_name", as_index=False)
        .agg(completion_rate=("completion_rate", "mean"), job_placement_rate=("job_placement_rate", "mean"))
        .sort_values("job_placement_rate", ascending=False)
    )
    funding_view = df.copy()
    funding_view["cost_per_participant"] = funding_view["funding_efficiency"].round(0)
    funding_view["estimated_placements"] = funding_view["placement_count"].round(0)

    template = "plotly_white"
    figures["participation_trends"] = px.line(
        trends,
        x="quarter",
        y="participants",
        markers=True,
        text="participants",
        title="TLD Workforce Participation by Reporting Quarter",
        labels={
            "quarter": "Reporting Quarter",
            "participants": "Participants Served",
        },
        hover_data={
            "participants": ":,",
            "workforce_completions": ":,.0f",
            "participant_growth": ":.1f",
            "trend_classification": True,
            "quarter": False,
        },
        color_discrete_sequence=[BLUE],
        template=template,
    )
    figures["placement_trends"] = px.line(
        trends,
        x="quarter",
        y="job_placement_rate",
        markers=True,
        text=trends["job_placement_rate"].round(1),
        title="Transportation, Logistics, and Distribution Placement Rate",
        labels={
            "quarter": "Reporting Quarter",
            "job_placement_rate": "Placement Rate (%)",
        },
        hover_data={
            "job_placement_rate": ":.1f",
            "credential_earned_rate": ":.1f",
            "employer_partnerships": ":,",
            "quarter": False,
        },
        color_discrete_sequence=[TEAL],
        template=template,
    )
    figures["top_programs"] = px.bar(
        top.sort_values("ranking_score"),
        x="ranking_score",
        y="program_name",
        orientation="h",
        color="program_label",
        text="ranking_score",
        title="Highest-Readiness Workforce Initiatives",
        labels={
            "ranking_score": "Composite Readiness Score",
            "program_name": "Workforce Initiative",
            "program_label": "Leadership Signal",
        },
        hover_data={
            "ranking_score": ":.1f",
            "workforce_readiness_score": ":.1f",
            "job_placement_rate": ":.1f",
            "credential_earned_rate": ":.1f",
            "participants": ":,",
            "logistics_specialization": True,
            "county": True,
        },
        template=template,
        color_discrete_map={
            "STRATEGIC ASSET": GREEN,
            "HIGH PERFORMING": BLUE,
            "NEEDS REVIEW": GOLD,
            "CRITICAL ATTENTION": RED,
        },
    )
    figures["bottom_programs"] = px.bar(
        bottom.sort_values("ranking_score", ascending=False),
        x="ranking_score",
        y="program_name",
        orientation="h",
        color="program_label",
        text="ranking_score",
        title="Initiatives Requiring Leadership Review",
        labels={
            "ranking_score": "Composite Readiness Score",
            "program_name": "Workforce Initiative",
            "program_label": "Leadership Signal",
        },
        hover_data={
            "ranking_score": ":.1f",
            "workforce_readiness_score": ":.1f",
            "job_placement_rate": ":.1f",
            "credential_earned_rate": ":.1f",
            "participants": ":,",
            "logistics_specialization": True,
            "county": True,
        },
        template=template,
        color_discrete_map={
            "STRATEGIC ASSET": GREEN,
            "HIGH PERFORMING": BLUE,
            "NEEDS REVIEW": GOLD,
            "CRITICAL ATTENTION": RED,
        },
    )
    figures["funding_vs_placement"] = px.scatter(
        funding_view,
        x="funding_amount",
        y="job_placement_rate",
        size="participants",
        color="logistics_specialization",
        hover_name="program_name",
        title="Funding Allocation vs. TLD Placement Outcomes",
        labels={
            "funding_amount": "Funding Allocation ($)",
            "job_placement_rate": "Placement Rate (%)",
            "participants": "Participants Served",
            "logistics_specialization": "TLD Focus Area",
            "cost_per_participant": "Cost per Participant ($)",
            "estimated_placements": "Estimated Placements",
            "county": "County",
            "quarter": "Quarter",
        },
        hover_data={
            "quarter": True,
            "county": True,
            "participants": ":,",
            "cost_per_participant": ":,.0f",
            "estimated_placements": ":,.0f",
            "credential_earned_rate": ":.1f",
        },
        template=template,
    )
    figures["specialization_distribution"] = px.pie(
        specialization,
        names="logistics_specialization",
        values="participants",
        title="Workforce Coverage by Port-Relevant TLD Focus Area",
        hole=0.38,
        color_discrete_sequence=px.colors.qualitative.Safe,
    )
    figures["readiness_heatmap"] = px.imshow(
        heatmap_data,
        aspect="auto",
        color_continuous_scale="Blues",
        title="Readiness Score by Initiative and Quarter",
        labels={"color": "Readiness Score", "x": "Reporting Quarter", "y": "Workforce Initiative"},
        text_auto=".0f",
    )
    figures["completion_vs_placement"] = px.bar(
        program_rates,
        x="program_name",
        y=["completion_rate", "job_placement_rate"],
        title="Completion and TLD Placement Rates by Initiative",
        barmode="group",
        text_auto=".1f",
        labels={
            "program_name": "Workforce Initiative",
            "value": "Rate (%)",
            "variable": "Performance Measure",
        },
        template=template,
        color_discrete_sequence=[BLUE, TEAL],
    )
    figures["county_participation"] = px.bar(
        county,
        x="county",
        y="participants",
        text="participants",
        title="Regional Workforce Participation by County",
        labels={"county": "County", "participants": "Participants Served"},
        color="participants",
        color_continuous_scale="Blues",
        template=template,
    )

    for fig in figures.values():
        fig.update_traces(hoverlabel=dict(bgcolor="white", font_size=13, font_family="Arial"))
        fig.update_layout(
            title_font_color=NAVY,
            font_family="Arial",
            margin=dict(l=30, r=30, t=70, b=55),
            legend_title_text="",
            hovermode="closest",
        )
    for key in ["top_programs", "bottom_programs", "completion_vs_placement", "county_participation"]:
        figures[key].update_traces(marker_line_width=1, marker_line_color="white", textposition="outside")
    for key in ["participation_trends", "placement_trends"]:
        figures[key].update_traces(marker=dict(size=11, line=dict(width=2, color="white")), textposition="top center")
    figures["funding_vs_placement"].update_traces(marker=dict(opacity=0.82, line=dict(width=1.5, color="white")))
    figures["completion_vs_placement"].update_xaxes(tickangle=35)
    figures["readiness_heatmap"].update_traces(hovertemplate="Initiative: %{y}<br>Quarter: %{x}<br>Readiness Score: %{z:.1f}<extra></extra>")
    return figures


def generate_executive_summary(df, trends=None, top_programs=None, bottom_programs=None):
    if trends is None:
        trends = calculate_trends(df)
    if top_programs is None:
        top_programs = identify_top_programs(df)
    if bottom_programs is None:
        bottom_programs = identify_bottom_programs(df)

    kpis = calculate_kpis(df)
    latest = trends.tail(1).iloc[0] if not trends.empty else None
    strongest_specialization = (
        df.groupby("logistics_specialization")["job_placement_rate"].mean().sort_values(ascending=False).index[0]
        if not df.empty
        else "transportation and logistics"
    )
    funding = calculate_funding_efficiency(df)
    efficient_program = funding.iloc[0]["program_name"] if not funding.empty else "N/A"
    watch_program = bottom_programs.iloc[0]["program_name"] if not bottom_programs.empty else "N/A"

    lines = [
        "Port Authority of New York & New Jersey - Port Department",
        "Port Policy & Planning Unit | Workforce Development Analytics Executive Summary",
        f"Generated: {datetime.now().strftime('%B %d, %Y')}",
        "",
        "Major Trends",
        f"- The active portfolio served {kpis.get('total_participants', 0):,} workforce participants across {kpis.get('active_programs', 0)} maritime, freight, and logistics-aligned programs.",
        f"- Average completion performance stands at {kpis.get('avg_completion_rate', 0)}%, with transportation-sector job placement averaging {kpis.get('avg_job_placement_rate', 0)}%.",
        f"- {strongest_specialization} initiatives currently show the strongest placement outcomes in the regional Transportation, Logistics, and Distribution labor pipeline.",
        "",
        "Operational Concerns",
        f"- {watch_program} should be reviewed for readiness, completion, placement, credential attainment, and participant retention performance.",
        "- Programs below target placement or credential attainment thresholds may limit the Port Department's ability to demonstrate measurable workforce readiness outcomes.",
        "- Any decline in employer partnerships should be treated as an early warning signal for training, career pathway, apprenticeship, and port-related placement capacity.",
        "",
        "Strategic Opportunities",
        f"- {efficient_program} provides a useful benchmark for cost-effective workforce delivery and should be examined for scalable practices.",
        "- Maritime logistics, terminal operations, trucking logistics, rail freight, and intermodal transportation programs should be compared against employer demand and Council on Port Performance workforce priorities.",
        "- Programs with strong completion but lower placement performance are candidates for deeper employer engagement and sector-specific credential alignment.",
        "",
        "Funding Insights",
        f"- Total workforce funding allocation in the filtered portfolio is ${kpis.get('total_funding', 0):,.0f}.",
        "- Leadership should review funding efficiency alongside placement outcomes rather than using participant volume alone as the primary performance signal.",
        "",
        "Transportation-Sector Employment Analysis",
        "- Placement outcomes indicate whether training investments are translating into port-related employment opportunities across cargo movement, warehousing, supply chain operations, and freight transportation roles.",
    ]

    if latest is not None:
        lines.extend(
            [
                f"- The latest reporting quarter, {latest['quarter']}, is classified as {latest['trend_classification']} for participant growth.",
                f"- Latest quarter workforce readiness averaged {latest['workforce_readiness_score']:.1f} out of 100.",
            ]
        )

    lines.extend(
        [
            "",
            "Recommendations for Leadership Review",
            "- Prioritize targeted review of programs labeled CRITICAL ATTENTION or NEEDS REVIEW before the next quarterly reporting cycle.",
            "- Expand employer partnership tracking for port-related placements, work-based learning, apprenticeships, and credential-bearing training tied to regional freight movement.",
            "- Use readiness score trends to brief Port Department leadership on workforce pipeline health and investment alignment.",
            "- Continue standardizing data collection across public-sector workforce partners to improve reporting speed and audit readiness.",
        ]
    )

    summary = "\n".join(lines)
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        SUMMARY_FILE.write_text(summary, encoding="utf-8")
    except Exception as exc:
        log_error(f"Failed to write executive summary: {exc}")
    return summary


def export_excel_dashboard(df, report_path=REPORT_FILE):
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        trends = calculate_trends(df)
        top = identify_top_programs(df)
        bottom = identify_bottom_programs(df)
        funding = calculate_funding_efficiency(df)
        kpis = pd.DataFrame([calculate_kpis(df)]).T.reset_index()
        kpis.columns = ["Metric", "Value"]
        insights = pd.DataFrame({"Strategic Insights": generate_executive_summary(df, trends, top, bottom).splitlines()})
        readiness = df[
            [
                "program_name",
                "quarter",
                "county",
                "logistics_specialization",
                "workforce_readiness_score",
                "program_label",
            ]
        ]

        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
            df.drop(columns=["quarter_sort"], errors="ignore").to_excel(writer, sheet_name="Cleaned Workforce Data", index=False)
            kpis.to_excel(writer, sheet_name="KPI Summary", index=False)
            trends.drop(columns=["quarter_sort"], errors="ignore").to_excel(writer, sheet_name="Trend Analysis", index=False)
            top.to_excel(writer, sheet_name="Top Programs", index=False)
            bottom.to_excel(writer, sheet_name="Bottom Programs", index=False)
            insights.to_excel(writer, sheet_name="Strategic Insights", index=False)
            funding.to_excel(writer, sheet_name="Funding Analysis", index=False)
            readiness.to_excel(writer, sheet_name="Workforce Readiness Scores", index=False)

        _format_workbook(report_path)
        return report_path
    except Exception as exc:
        log_error(f"Export failure: {exc}")
        return None


def _format_workbook(report_path):
    try:
        workbook = load_workbook(report_path)
        header_fill = PatternFill("solid", fgColor="E9F2F8")
        header_font = Font(color="244A73", bold=True)
        thin_border = Border(bottom=Side(style="thin", color="D9E2EC"))
        warning_fill = PatternFill("solid", fgColor="FCE8E6")

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 14), 44)
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for header_cell in worksheet[1]:
                if header_cell.value in {
                    "completion_rate",
                    "job_placement_rate",
                    "credential_earned_rate",
                    "workforce_readiness_score",
                    "ranking_score",
                }:
                    col = header_cell.column_letter
                    worksheet.conditional_formatting.add(
                        f"{col}2:{col}{worksheet.max_row}",
                        CellIsRule(operator="lessThan", formula=["70"], fill=warning_fill),
                    )
        workbook.save(report_path)
    except Exception as exc:
        log_error(f"Formatting issue: {exc}")


def _quarter_sort_key(value):
    match = re.match(r"^Q([1-4])\s(20\d{2})$", str(value).strip())
    if not match:
        return 0
    quarter, year = match.groups()
    return int(year) * 10 + int(quarter)


def _participant_growth_score(df):
    working = df.copy()
    working["participant_growth_score"] = 50.0
    for program_name, group in working.sort_values("quarter_sort").groupby("program_name"):
        pct = group["participants"].pct_change().replace([np.inf, -np.inf], 0).fillna(0)
        score = (50 + pct * 100).clip(0, 100)
        working.loc[group.index, "participant_growth_score"] = score
    return working["participant_growth_score"]


def _workforce_readiness(df):
    placement_volume_score = np.where(
        df["participants"] > 0,
        (df["port_related_job_placements"] / df["participants"] * 100).clip(0, 100),
        0,
    )
    score = (
        df["completion_rate"] * 0.30
        + df["job_placement_rate"] * 0.35
        + df["credential_earned_rate"] * 0.15
        + placement_volume_score * 0.10
        + df["participant_growth_score"] * 0.10
    )
    return score.clip(0, 100).round(1)


def _program_rankings(df):
    program = (
        df.groupby("program_name", as_index=False)
        .agg(
            workforce_readiness_score=("workforce_readiness_score", "mean"),
            funding_efficiency=("funding_efficiency", "mean"),
            job_placement_rate=("job_placement_rate", "mean"),
            participant_retention=("participant_retention", "mean"),
            credential_earned_rate=("credential_earned_rate", "mean"),
            port_related_job_placements=("port_related_job_placements", "sum"),
            participants=("participants", "sum"),
            logistics_specialization=("logistics_specialization", "first"),
            county=("county", "first"),
        )
    )
    efficiency_score = 100 - _normalize(program["funding_efficiency"])
    placement_volume_score = np.where(
        program["participants"] > 0,
        (program["port_related_job_placements"] / program["participants"] * 100).clip(0, 100),
        0,
    )
    program["ranking_score"] = (
        program["workforce_readiness_score"] * 0.35
        + efficiency_score * 0.15
        + program["job_placement_rate"] * 0.20
        + program["participant_retention"] * 0.10
        + program["credential_earned_rate"] * 0.10
        + placement_volume_score * 0.10
    ).round(1)
    program["program_label"] = program["ranking_score"].apply(_score_label)
    return program.sort_values("ranking_score", ascending=False).reset_index(drop=True)


def _normalize(series):
    minimum = series.min()
    maximum = series.max()
    if maximum == minimum:
        return pd.Series(50, index=series.index)
    return ((series - minimum) / (maximum - minimum) * 100).clip(0, 100)


def _score_label(score):
    if score >= 85:
        return "STRATEGIC ASSET"
    if score >= 75:
        return "HIGH PERFORMING"
    if score >= 60:
        return "NEEDS REVIEW"
    return "CRITICAL ATTENTION"


def _program_label(row):
    return _score_label(row["workforce_readiness_score"])


def _classify_trend(value):
    if value >= 20:
        return "High Growth"
    if value >= 5:
        return "Improving"
    if value <= -15:
        return "Critical Decline"
    if value <= -5:
        return "Declining"
    return "Stable"


def _rename_for_display(df, columns):
    if df.empty:
        return df
    display_df = df.copy()
    available = [column for column in columns if column in display_df.columns]
    display_df = display_df[available]
    return display_df.rename(columns=columns)


def _program_display_columns():
    return {
        "program_name": "Workforce Initiative",
        "workforce_readiness_score": "Readiness Score",
        "funding_efficiency": "Cost per Participant ($)",
        "job_placement_rate": "TLD Placement Rate (%)",
        "participant_retention": "Completion / Retention (%)",
        "credential_earned_rate": "Credential Attainment (%)",
        "port_related_job_placements": "Port-Related Placements",
        "participants": "Participants Served",
        "logistics_specialization": "TLD Focus Area",
        "county": "County",
        "ranking_score": "Composite Score",
        "program_label": "Leadership Signal",
    }
